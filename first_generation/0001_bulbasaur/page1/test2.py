import openpyxl

game_dict = {
    'Generation I': ['Game', 'I'],
    'Pokémon Red and Blue': ['Game', 'I'],
    'original series': ['Anime', 'I/II'],
    'Pokémon Yellow': ['Game', 'I'],
    'Pokémon Crystal': ['Game', 'II'],
    'Pokémon Crystal': ['Game', 'II'],
    'Pokémon Gold': ['Game', 'II'],
    'Pokémon Silver': ['Game', 'II'],
    'Pokémon Gold and Silver': ['Game', 'II'],
    'Pokémon Colosseum': ['Game', 'III'],
    'Pokémon XD: Gale of Darkness': ['Game', 'III'],
    'Pokémon Channel': ['Game', 'III'],
    'Pokémon FireRed and LeafGreen': ['Game', 'III'],
    'Pokémon Ruby and Sapphire': ['Game', 'III'],
    'Generation II': ['Game', 'II'],
    'Generation III': ['Game', 'III'],
    'Generation IV': ['Game', 'IV'],
    'Generation V': ['Game', 'V'],
    'Pokémon the Series: Ruby and Sapphire': ['Anime', 'III'],
    'Pokémon HeartGold and SoulSilver': ['Game', 'IV'],
    'Pokémon Battle Revolution': ['Game', 'IV'],
    'Pokémon Diamond and Pearl': ['Game', 'IV'],
    'Pokémon Platinum': ['Game', 'IV'],
    'Pokémon Black and White': ['Game', 'V'],
    'Pokémon Black and White 2': ['Game', 'V'],
    'Pokémon Black 2 and White 2': ['Game', 'V'],
    'Pokédex 3D Pro': ['App', 'V'],
    'Pokémon Global Link': ['Game', 'V'],
    'Nintendo Badge Arcade': ['Game', 'V'],
    'Pokémon Battle Trozei': ['Game', 'VI'],
    'Pokémon Super Mystery Dungeon': ['Game', 'VI'],
    'Pokémon the Series: XY': ['Anime', 'VI'],
    'Pokémon Omega Ruby and Alpha Sapphire': ['Game', 'VI'],
    'Pokémon Sun and Moon': ['Game', 'VII'],
    "Pokémon: Let's Go, Pikachu! and Let's Go, Eevee!": ['Game', 'VII'],
    "Pokémon Brilliant Diamond and Shining Pearl": ['Game', 'VIII'],
    "Pokémon Café ReMix": ['App', 'VIII'],
    "Pokémon Mystery Dungeon: Rescue Team DX": ['Game', 'VIII'],
    "Pokémon Smile": ['App', 'VIII'],
    "Pokémon Sword and Shield": ['Game', 'VIII'],
    "Pokémon: Kids Winter Fest": ['Website', 'VIII'],
}

types = ["back sprite", "official artwork", "official art", "game sprite", "artwork", "menusprite", "boxsprite", "sprite", "icon", "back model", "model"]

# Open the text file in read mode
with open('0001_catalogue.txt', 'r', encoding="utf-8") as text_file:
    # Create a new Excel workbook and select the default sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Define your header row as a list of values
    header = ["Filename", "Type", "Origin Type", "Origin", "Generation", "Is Shiny"]

    # Write the header row to the first row of the sheet
    sheet.append(header)

    # Iterate through each line in the text file and split it into two columns
    for row_index, line in enumerate(text_file, start=1):
        if '.png' in line:
            name, extra = line.split('.png: ')
        else:
            # If '.png' is not found in the string, you can choose to handle it accordingly
            name = line
        
        origin = ""
        origin_types= ""
        generations = ""
        type = ""
        
        for key, value in game_dict.items():
            if key in extra:
                origin += key + '/'
                origin_types += value[0] + '/'
                generations += value[1] + '/'
        
        for value in types:
            if value in extra.lower():
                type = value

        sheet.cell(row=row_index+1, column=1, value=name.strip())
        sheet.cell(row=row_index+1, column=2, value=type.strip('/'))
        sheet.cell(row=row_index+1, column=3, value=origin_types.strip('/'))
        sheet.cell(row=row_index+1, column=4, value=origin.strip('/'))
        sheet.cell(row=row_index+1, column=5, value=generations.strip('/'))



    # Save the Excel file with a desired name
    workbook.save('0001_catalogue.xlsx')